"""
Export Service - Phase 2A Step 2
Handles PDF and Excel export functionality
Location: C:\service-gateway\services\export_service.py
"""

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting query results to PDF and Excel"""
    
    def __init__(self):
        self.company_name = "Förlagssystem AB"
        self.company_color = colors.HexColor("#0073AE")
        
        # Import column mapping from RBAC rules
        from utils.rbac_rules import COLUMN_FRIENDLY_NAMES
        self.column_mapping = COLUMN_FRIENDLY_NAMES
    
    def _get_friendly_column_name(self, technical_name: str) -> str:
        """Convert technical column name to user-friendly name"""
        return self.column_mapping.get(technical_name, technical_name)
    
    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str,
        user_name: str,
        query: str = None
    ) -> BytesIO:
        """
        Export data to PDF format
        
        Args:
            data: List of dictionaries containing query results
            title: Report title
            user_name: Name of user requesting export
            query: Original query text (optional)
            
        Returns:
            BytesIO buffer containing PDF
        """
        buffer = BytesIO()
        
        try:
            # Create PDF document
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18,
            )
            
            # Container for PDF elements
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=self.company_color,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=self.company_color,
                spaceAfter=12,
            )
            
            # Header
            elements.append(Paragraph(self.company_name, title_style))
            elements.append(Spacer(1, 12))
            
            # Report info
            elements.append(Paragraph(title, heading_style))
            
            info_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
            info_text += f"Requested by: {user_name}<br/>"
            info_text += f"Records: {len(data)}"
            
            elements.append(Paragraph(info_text, styles['Normal']))
            elements.append(Spacer(1, 20))
            
            # Add query if provided
            if query:
                elements.append(Paragraph("Query:", heading_style))
                elements.append(Paragraph(query, styles['Code']))
                elements.append(Spacer(1, 20))
            
            # Data table
            if data and len(data) > 0:
                # Get column names and map to friendly names
                technical_columns = list(data[0].keys())
                friendly_columns = [self._get_friendly_column_name(col) for col in technical_columns]
                
                # Prepare table data with friendly headers
                table_data = [friendly_columns]  # Header row
                
                for row in data:
                    table_data.append([str(row.get(col, '')) for col in technical_columns])
                
                # Calculate column widths
                available_width = doc.width
                col_width = available_width / len(technical_columns)
                
                # Create table
                table = Table(table_data, colWidths=[col_width] * len(technical_columns))
                
                # Table style
                table.setStyle(TableStyle([
                    # Header
                    ('BACKGROUND', (0, 0), (-1, 0), self.company_color),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    
                    # Body
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                    
                    # Grid
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ]))
                
                elements.append(table)
            else:
                elements.append(Paragraph("No data to display", styles['Normal']))
            
            # Footer
            elements.append(Spacer(1, 30))
            footer_text = f"{self.company_name} - Confidential"
            elements.append(Paragraph(footer_text, styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            
            buffer.seek(0)
            logger.info(f"PDF generated successfully: {len(data)} rows")
            return buffer
            
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            raise
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        title: str,
        user_name: str,
        query: str = None
    ) -> BytesIO:
        """
        Export data to Excel format
        
        Args:
            data: List of dictionaries containing query results
            title: Report title
            user_name: Name of user requesting export
            query: Original query text (optional)
            
        Returns:
            BytesIO buffer containing Excel file
        """
        buffer = BytesIO()
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Query Results"
            
            # Styles
            header_fill = PatternFill(start_color="0073AE", end_color="0073AE", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=14, color="0073AE")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title and metadata
            ws['A1'] = self.company_name
            ws['A1'].font = title_font
            
            ws['A2'] = title
            ws['A2'].font = Font(bold=True, size=12)
            
            ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A4'] = f"Requested by: {user_name}"
            ws['A5'] = f"Records: {len(data)}"
            
            # Add query if provided
            start_row = 7
            if query:
                ws[f'A{start_row}'] = "Query:"
                ws[f'A{start_row}'].font = Font(bold=True)
                ws[f'A{start_row + 1}'] = query
                start_row += 3
            
            # Data
            if data and len(data) > 0:
                # Get column names and map to friendly names
                technical_columns = list(data[0].keys())
                friendly_columns = [self._get_friendly_column_name(col) for col in technical_columns]
                
                # Header row with friendly names
                for col_idx, column in enumerate(friendly_columns, start=1):
                    cell = ws.cell(row=start_row, column=col_idx)
                    cell.value = column
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = border
                
                # Data rows
                for row_idx, row_data in enumerate(data, start=start_row + 1):
                    for col_idx, tech_column in enumerate(technical_columns, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = row_data.get(tech_column, '')
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='top')
                        
                        # Alternate row colors
                        if row_idx % 2 == 0:
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                # Auto-adjust column widths based on friendly column names
                for col_idx, column in enumerate(friendly_columns, start=1):
                    max_length = len(str(column))
                    for row_idx in range(start_row + 1, start_row + len(data) + 1):
                        cell_value = str(ws.cell(row=row_idx, column=col_idx).value)
                        max_length = max(max_length, len(cell_value))
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width
            
            # Save to buffer
            wb.save(buffer)
            buffer.seek(0)
            
            logger.info(f"Excel generated successfully: {len(data)} rows")
            return buffer
            
        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            raise